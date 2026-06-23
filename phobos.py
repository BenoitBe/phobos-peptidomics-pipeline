# ==============================================================================
# phobos.py — Orchestrateur principal
# Phobos — Peptide-level Heuristics for Bottom-up Omics Suite
# Peptide-centric DDA Proteomics Pipeline (Peaks export)
# ==============================================================================
# Compagnon de Deimos pour les données DDA (Peaks Studio).
# Niveau d'analyse : peptide unique (séquence + charge + modifications).
#
# Différences structurelles vs deimos.py :
#   • Input     : protein-peptides.csv (Peaks, export combiné)
#   • Intensité : Area brute → normalisation médiane interne
#   • Filtrage  : sur score Peaks (-10lgP) + validité FDR peptide
#   • Unité     : peptide (Peptide + Charge + Modifications = clé unique)
#   • Modèle    : limma eBayes sur log2(Area) normalisé
#   • Agrégation protéine optionnelle (median polish post-hoc, visualisation)
#   • Pas de DEqMS (on EST au niveau peptide)
#   • Pas de GO (pas demandé pour l'instant)
#
# Modules réutilisés depuis Deimos :
#   limma_ebayes.py (lm_fit, contrasts_fit, ebayes, top_table,
#                    make_all_contrasts, make_design_matrix, _bh_correction)
#   config.py       (resolve_config, YAML I/O)
#
# Dépendances :
#   pandas, numpy, scipy, statsmodels, sklearn, umap-learn,
#   matplotlib, seaborn, openpyxl, pillow, PyComplexHeatmap (optionnel),
#   adjustText (optionnel), pyyaml
# ==============================================================================

import os
import sys
import re
import warnings
import logging
import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from matplotlib.backends.backend_pdf import PdfPages
from matplotlib.colors import to_hex
import seaborn as sns
from math import ceil
from itertools import combinations
from scipy import stats
from scipy.stats import norm, pearsonr
from scipy.cluster.hierarchy import linkage, dendrogram, fcluster
from scipy.spatial.distance import squareform
from sklearn.decomposition import PCA
from sklearn.preprocessing import StandardScaler
from sklearn.impute import SimpleImputer
import umap
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Alignment
import openpyxl

warnings.filterwarnings("ignore")
logging.getLogger("adjustText").setLevel(logging.ERROR)

# Modules Deimos réutilisés (doivent être dans le même dossier)
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from limma_ebayes import (lm_fit, contrasts_fit, ebayes, top_table,
                           make_all_contrasts, make_design_matrix,
                           _bh_correction)
from config import resolve_config


# ==============================================================================
# 0. PARAMÈTRES INTERACTIFS
# ==============================================================================

def ask_params() -> dict:
    print("\n" + "="*60)
    print("  PHOBOS — DDA Peptide Pipeline — Threshold configuration")
    print("="*60)

    # Volcanos / robustness
    print("\n[1/2] Individual volcanos & robustness score")
    v_type = ""
    while v_type not in ("1", "2"):
        v_type = input("  Use (1) raw p.value  or  (2) p.adj / FDR [BH] ? -> ").strip()
    v_val = None
    while v_val is None:
        try:
            v_val = float(input("  Threshold value (e.g. 0.05) -> ").strip())
        except ValueError:
            pass
    v_ratio = None
    while v_ratio is None:
        try:
            v_ratio = float(input("  Minimum ratio (e.g. 1.5, 2.0) -> ").strip())
        except ValueError:
            pass

    # ANOVA / Heatmap
    print("\n[2/2] ANOVA & heatmaps")
    a_type = ""
    while a_type not in ("1", "2"):
        a_type = input("  Use (1) raw p.value  or  (2) p.adj / FDR [BH] ? -> ").strip()
    a_val = None
    while a_val is None:
        try:
            a_val = float(input("  Threshold value -> ").strip())
        except ValueError:
            pass

    # Filtrage Peaks
    print("\n[Peaks QC] Peptide score filtering")
    score_thr = None
    while score_thr is None:
        raw = input("  Minimum -10lgP score [default 20] -> ").strip()
        score_thr = 20.0 if raw == "" else (float(raw) if raw.replace(".", "").isdigit() else None)

    # Clusters heatmap
    n_clusters = None
    while n_clusters is None:
        raw = input("  Number of clusters (heatmap) [default 3] -> ").strip()
        if raw == "":
            n_clusters = 3
        else:
            try:
                n_clusters = int(raw)
                if n_clusters < 2:
                    n_clusters = None
            except ValueError:
                pass

    # Robustness
    n_iter_rob = None
    while n_iter_rob is None:
        raw = input("  Robustness iterations [default 100, 0=disabled] -> ").strip()
        if raw == "":
            n_iter_rob = 100
        else:
            try:
                n_iter_rob = int(raw)
                if n_iter_rob < 0:
                    n_iter_rob = None
            except ValueError:
                pass

    # FDR global vs par contraste
    fdr_global = None
    while fdr_global is None:
        raw = input("  FDR correction: [1] per contrast  [2] global -> ").strip()
        if raw in ("", "1"):
            fdr_global = False
        elif raw == "2":
            fdr_global = True

    # Dashboard HTML interactif
    raw = input("  Generate the interactive HTML dashboard? [Y/n] -> ").strip().lower()
    make_dashboard = raw in ("", "y", "yes", "o", "oui")

    params = {
        "volcano_use_padj":   v_type == "2",
        "volcano_p_thresh":   v_val,
        "volcano_lfc_min":    np.log2(v_ratio),
        "volcano_ratio_min":  v_ratio,
        "anova_use_padj":     a_type == "2",
        "anova_p_thresh":     a_val,
        "n_heatmap_clusters": n_clusters,
        "n_iter_robustness":  n_iter_rob,
        "fdr_global":         fdr_global,
        "peaks_score_thr":    score_thr,
        "impute_method":      "qrilc",   # fixé QRILC (DDA MNAR dominant)
        "go_organism":        None,
        "make_dashboard":     make_dashboard,
        "use_deqms":          False,
    }

    print(f"\n  [OK] Volcanos -> "
          f"{'p.adj' if params['volcano_use_padj'] else 'p.value'} < {v_val} "
          f"| ratio >= {v_ratio}")
    print(f"       ANOVA   -> "
          f"{'p.adj' if params['anova_use_padj'] else 'p.value'} < {a_val}")
    print(f"       Peaks score thr -> {score_thr} | Clusters -> {n_clusters} "
          f"| Robustness -> {n_iter_rob} iter")
    return params


# ==============================================================================
# 1. LECTURE ET PARSING DU protein-peptides.csv (Peaks)
# ==============================================================================

# Colonnes méta attendues dans le fichier Peaks protein-peptides.csv.
# La structure réelle varie légèrement selon la version de Peaks et
# les options d'export — le parser s'adapte à ce qui est présent.
_PEAKS_PEPTIDE_META = [
    "Peptide", "Modifications", "Charge", "m/z", "Mass", "Accession",
    "-10lgP", "ppm", "RT", "Scan", "Source File",
]
_PEAKS_PROTEIN_META = [
    "Protein Group", "Protein ID", "Description",
]


def load_peaks_data(csv_path: str, design_path: str,
                    score_thr: float = 20.0
                    ) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, list]:
    """
    Lit le protein-peptides.csv de Peaks et le design expérimental.

    Format Peaks : le fichier contient à la fois des lignes « protéine »
    (en gras dans l'interface) et des lignes « peptide ». Dans l'export CSV,
    les lignes protéine ont Peptide == NaN (ou une syntaxe différente) et
    les lignes peptide ont des valeurs numériques dans les colonnes d'Area.

    Returns
    -------
    df_pep  : DataFrame peptides filtrés + colonnes Area log2 normalisées
    meta    : colonnes d'annotation peptide (séquence, modifs, protéine…)
    design  : design expérimental aligné sur les colonnes Area
    area_cols : noms de colonnes Area (un par échantillon)
    """
    print(f"[IO] Reading Peaks peptide export: {csv_path}")
    ext = os.path.splitext(csv_path)[1].lower()
    if ext in (".xlsx", ".xlsm", ".xls"):
        # Export .xlsx (TIMS-TOF / Peaks) — feuille 'Peptides' si présente
        xl = pd.ExcelFile(csv_path)
        sheet = "Peptides" if "Peptides" in xl.sheet_names else xl.sheet_names[0]
        raw = pd.read_excel(csv_path, sheet_name=sheet)
        print(f"  -> Excel sheet '{sheet}' loaded")
    else:
        raw = pd.read_csv(csv_path, sep=",", low_memory=False)
    raw.columns = [str(c).strip() for c in raw.columns]
    print(f"  -> {len(raw)} rows, {len(raw.columns)} columns")

    # ── Identifier les colonnes Area (une par run/échantillon) ──────────────
    # Peaks nomme ces colonnes "Area" (si un seul run) ou "Area <sample_name>"
    area_cols = [c for c in raw.columns
                 if re.match(r"^Area\b", c, re.IGNORECASE)]
    if not area_cols:
        # Repli : colonnes dont le nom contient "area" (case-insensitive)
        area_cols = [c for c in raw.columns if "area" in c.lower()]
    if not area_cols:
        raise ValueError(
            "[ERROR] No 'Area' column found in the Peaks CSV. "
            "Check the export format (protein-peptides.csv).")
    print(f"  -> {len(area_cols)} Area columns detected: {area_cols[:6]}{'...' if len(area_cols)>6 else ''}")

    # ── Garder uniquement les lignes peptide ─────────────────────────────────
    # Les lignes protéines sont identifiées par l'absence de valeur dans
    # la colonne "Peptide" (ou un marqueur spécifique selon la version Peaks).
    pep_col = next((c for c in raw.columns
                    if c.strip().lower() in ("peptide", "sequence")), None)
    if pep_col is None:
        raise ValueError("[ERROR] 'Peptide' column not found. Check the CSV header.")

    df = raw.dropna(subset=[pep_col]).copy()
    # Exclure les lignes dont la valeur dans 'Peptide' ressemble à une protéine
    # (certaines versions Peaks insèrent un header de groupe dans cette colonne)
    df = df[~df[pep_col].astype(str).str.startswith("Protein")].reset_index(drop=True)
    print(f"  -> {len(df)} peptide rows retained (protein header rows removed)")

    # ── Filtrage sur le score Peaks (-10lgP) ────────────────────────────────
    score_col = next((c for c in df.columns if "-10lgp" in c.lower()), None)
    if score_col and score_thr is not None:
        n_before = len(df)
        df = df[pd.to_numeric(df[score_col], errors="coerce") >= score_thr].reset_index(drop=True)
        print(f"  -> Score filter (-10lgP >= {score_thr}): {n_before} -> {len(df)} peptides")

    # ── Construction de l'identifiant peptide unique ─────────────────────────
    # Peaks : même séquence peut exister avec charges et/ou modifs différentes
    # → on les traite comme des features DISTINCTES (niveau précurseur).
    # NB : dans l'export TIMS-TOF/.xlsx, les modifs sont (a) inline dans la
    # séquence sous forme (+x.xx) ET (b) résumées dans la colonne 'PTM'.
    mod_col = next((c for c in df.columns
                    if c.strip().lower() in ("modifications", "modification",
                                             "variable modifications", "ptm")), None)
    charge_col = next((c for c in df.columns
                       if c.strip().lower() in ("charge", "z")), None)

    df["_seq"] = df[pep_col].astype(str).str.strip()
    df["_mod"] = df[mod_col].fillna("").astype(str).str.strip() if mod_col else ""
    df["_chg"] = df[charge_col].fillna("").astype(str).str.strip() if charge_col else ""

    # Séquence "stripped" : retire les masses delta inline (+0.98), (-17.03)…
    # et tout caractère non amino-acide → backbone nu pour regroupement/longueur.
    df["_seq_stripped"] = (df["_seq"]
                           .str.replace(r"\([+-]?\d+\.?\d*\)", "", regex=True)
                           .str.replace(r"[^A-Z]", "", regex=True))

    # peptide_id = séquence AVEC modif inline (distinction native modifié/non),
    # complétée par la charge pour séparer les états de charge.
    df["peptide_id"] = df["_seq"] + "|z=" + df["_chg"]

    # Dédoublonnage : si plusieurs lignes pour le même peptide_id (cas rare
    # avec les exports groupés), on garde celle avec le meilleur score,
    # ou on fait la somme des aires si le score est absent.
    if df["peptide_id"].duplicated().any():
        if score_col:
            df[score_col] = pd.to_numeric(df[score_col], errors="coerce")
            df = (df.sort_values(score_col, ascending=False)
                    .drop_duplicates("peptide_id")
                    .reset_index(drop=True))
        else:
            # Agrégation par somme sur les colonnes Area (somme des coélutions)
            num_cols = area_cols
            agg = {c: "sum" for c in num_cols}
            first_cols = {c: "first" for c in df.columns if c not in num_cols + ["peptide_id"]}
            agg.update(first_cols)
            df = df.groupby("peptide_id", as_index=False).agg(agg)
        print(f"  -> After deduplication: {len(df)} unique peptide features")

    # ── Colonnes méta peptide ────────────────────────────────────────────────
    # On récupère les colonnes protéine si présentes (Accession, Description…)
    acc_col = next((c for c in df.columns
                    if c.strip().lower() in ("accession", "protein group",
                                             "protein id", "protein")), None)
    desc_col = next((c for c in df.columns
                     if "description" in c.lower()), None)
    gene_col = next((c for c in df.columns
                     if c.strip().lower() in ("gene", "genes", "gene name")), None)
    rt_col  = next((c for c in df.columns if "rt" in c.strip().lower()), None)
    mz_col  = next((c for c in df.columns if c.strip().lower() in ("m/z", "mz")), None)

    meta_keep = ["peptide_id", "_seq", "_seq_stripped", "_mod", "_chg"]
    for c in [acc_col, desc_col, gene_col, score_col, rt_col, mz_col]:
        if c and c not in meta_keep:
            meta_keep.append(c)
    meta_keep = [c for c in meta_keep if c in df.columns]
    meta = df[meta_keep].copy().reset_index(drop=True)
    meta.rename(columns={
        "_seq": "Sequence", "_seq_stripped": "Sequence_stripped",
        "_mod": "Modifications", "_chg": "Charge",
        acc_col: "Accession", desc_col: "Description",
        gene_col: "Gene", score_col: "Score_10lgP",
    }, errors="ignore", inplace=True)

    # ── Design ───────────────────────────────────────────────────────────────
    design = pd.read_csv(design_path, sep=";")
    print(f"  -> Design: {len(design)} samples, columns: {list(design.columns)}")

    # Aligner colonnes Area sur le design
    area_cols_aligned, design_aligned = _align_area_to_design(
        area_cols, design, df)

    if not area_cols_aligned:
        raise ValueError(
            "[ERROR] No Area column could be matched to the design labels. "
            "Check that 'label' entries in the design match the Area column names.")

    # ── Matrice intensités ───────────────────────────────────────────────────
    mat = df[area_cols_aligned].copy().astype(float)
    mat.replace(0, np.nan, inplace=True)

    return df, meta, mat, design_aligned, area_cols_aligned


def _align_area_to_design(area_cols: list, design: pd.DataFrame,
                           df: pd.DataFrame
                           ) -> tuple[list, pd.DataFrame]:
    """
    Fait correspondre les colonnes Area aux labels du design.

    Convention (Option B) : le 'label' du design est le SUFFIXE de la colonne
    Area. Pour un label 'MA1', le parser cherche en priorité la colonne
    'Area MA1' (préfixe 'Area ' + label EXACT). Cette correspondance stricte
    lève l'ambiguïté classique (ex. label 'S1' ne doit pas matcher 'Area MS1').

    Stratégie multi-pass, du plus strict au plus permissif :
      1. 'Area <label>'  exact (insensible casse/espaces multiples)
      2. label == nom complet de colonne (Option A tolérée)
      3. label == colonne sans le préfixe 'Area ' (variantes d'espacement)
      4. Levenshtein <= 2 sur le suffixe (repli prudent, seuil resserré)

    En cas de collision (deux labels matchant la même colonne, ou un label
    matchant plusieurs colonnes), un avertissement explicite est émis.
    """
    def _norm(s: str) -> str:
        # normalise espaces multiples et casse pour comparaison robuste
        return re.sub(r"\s+", " ", str(s).strip()).lower()

    # Index normalisé des colonnes Area et de leur suffixe (sans 'Area ')
    col_norm   = {c: _norm(c) for c in area_cols}
    col_suffix = {c: _norm(re.sub(r"^area\s+", "", c, flags=re.IGNORECASE))
                  for c in area_cols}

    matched_cols, matched_rows = [], []
    used_cols = set()

    for _, row in design.iterrows():
        lbl = str(row["label"]).strip()
        lbl_n = _norm(lbl)
        target_exact = _norm(f"Area {lbl}")
        hit = None

        # Pass 1 : 'Area <label>' exact
        for c in area_cols:
            if c in used_cols:
                continue
            if col_norm[c] == target_exact:
                hit = c; break
        # Pass 2 : label == nom complet de colonne
        if hit is None:
            for c in area_cols:
                if c not in used_cols and col_norm[c] == lbl_n:
                    hit = c; break
        # Pass 3 : label == suffixe (sans 'Area ')
        if hit is None:
            for c in area_cols:
                if c not in used_cols and col_suffix[c] == lbl_n:
                    hit = c; break
        # Pass 4 : Levenshtein <= 2 sur le suffixe (repli resserré)
        if hit is None:
            best_d, best_c = 999, None
            for c in area_cols:
                if c in used_cols:
                    continue
                d = _levenshtein(lbl_n, col_suffix[c])
                if d < best_d:
                    best_d, best_c = d, c
            if best_d <= 2:
                hit = best_c
                print(f"  [WARN] Label '{lbl}' matched to '{hit}' by fuzzy "
                      f"distance ({best_d}) — verify this is correct.")

        if hit:
            matched_cols.append(hit)
            matched_rows.append(row)
            used_cols.add(hit)
        else:
            print(f"  [WARN] Label '{lbl}' not matched to any Area column "
                  f"— sample excluded")

    design_out = pd.DataFrame(matched_rows).reset_index(drop=True)
    n_design = len(design)
    n_match  = len(matched_cols)
    if n_match < n_design:
        print(f"  [WARN] {n_design - n_match}/{n_design} design label(s) "
              f"unmatched. Check the 'label' column against the Area headers.")
    return matched_cols, design_out


def _levenshtein(a: str, b: str) -> int:
    if len(a) < len(b):
        return _levenshtein(b, a)
    if not b:
        return len(a)
    prev = list(range(len(b) + 1))
    for ca in a:
        curr = [prev[0] + 1]
        for j, cb in enumerate(b):
            curr.append(min(prev[j+1]+1, curr[j]+1, prev[j]+(ca != cb)))
        prev = curr
    return prev[-1]


def diagnose_labels_peaks(csv_path: str, design_path: str) -> bool:
    """Diagnostic de cohérence colonnes Area ↔ design (équivalent de deimos.diagnose_labels)."""
    print("\n" + "="*60)
    print("  DIAGNOSTIC — Peaks Area columns vs Design")
    print("="*60)
    ok = True
    for p in [csv_path, design_path]:
        if not os.path.exists(p):
            print(f"  [ERROR] File not found: {p}")
            ok = False
    if not ok:
        return False

    design = pd.read_csv(design_path, sep=";")
    if "label" not in design.columns:
        print("  [ERROR] 'label' column missing in the design")
        return False

    ext = os.path.splitext(csv_path)[1].lower()
    if ext in (".xlsx", ".xlsm", ".xls"):
        xl = pd.ExcelFile(csv_path)
        sheet = "Peptides" if "Peptides" in xl.sheet_names else xl.sheet_names[0]
        raw = pd.read_excel(csv_path, sheet_name=sheet, nrows=0)
    else:
        raw = pd.read_csv(csv_path, sep=",", nrows=0)  # header only
    raw.columns = [str(c).strip() for c in raw.columns]
    area_cols = [c for c in raw.columns if re.match(r"^Area\b", c, re.IGNORECASE)]
    if not area_cols:
        area_cols = [c for c in raw.columns if "area" in c.lower()]

    def _norm(s):
        return re.sub(r"\s+", " ", str(s).strip()).lower()
    col_norm   = {c: _norm(c) for c in area_cols}
    col_suffix = {c: _norm(re.sub(r"^area\s+", "", c, flags=re.IGNORECASE))
                  for c in area_cols}

    design_labels = design["label"].astype(str).tolist()

    print(f"\n  Area columns  : {len(area_cols)}")
    print(f"  Design labels : {len(design_labels)}")

    matched, unmatched_d = [], []
    matched_cols = set()
    for lbl in design_labels:
        lbl_n = _norm(lbl)
        target = _norm(f"Area {lbl}")
        hit = None
        for c in area_cols:
            if col_norm[c] == target or col_norm[c] == lbl_n or col_suffix[c] == lbl_n:
                hit = c; break
        if hit:
            matched.append(lbl); matched_cols.add(hit)
        else:
            unmatched_d.append(lbl)

    unmatched_a = [col_suffix[c] for c in area_cols if c not in matched_cols]

    print(f"  [OK] Matched : {len(matched)}")
    if unmatched_d:
        print(f"  [ERROR] In design but not matched to any Area col ({len(unmatched_d)}):")
        for l in unmatched_d:
            print(f"     - '{l}'  (expected column 'Area {l}')")
        ok = False
    if unmatched_a:
        print(f"  [WARN] Area columns not covered by design ({len(unmatched_a)}):")
        for s in unmatched_a:
            print(f"     - '{s}'")
    if ok:
        print(f"\n  [OK] Labels consistent — {len(matched)} samples ready.\n")
    else:
        print("\n  [STOP] Fix the design or column names before re-running.\n")
    return ok


# ==============================================================================
# 2. LOG2 + NORMALISATION MÉDIANE
# ==============================================================================

def log2_normalize(mat: pd.DataFrame) -> pd.DataFrame:
    """
    log2(Area) puis normalisation médiane par échantillon.
    La normalisation médiane centre chaque run sur la médiane globale
    (médiane des médianes) sans supposer une distribution spécifique.
    Équivalent de normalizeMedianValues() de DEP/R.
    """
    mat_log = np.log2(mat.replace(0, np.nan))

    # Médiane par échantillon (NaN ignorés)
    col_medians = mat_log.median(axis=0)       # (n_samples,)
    grand_median = col_medians.median()        # scalaire

    # Décalage pour que chaque run ait la même médiane
    shift = grand_median - col_medians         # (n_samples,)
    mat_norm = mat_log.add(shift, axis=1)

    print(f"  [OK] Median normalization: grand median = {grand_median:.3f} log2")
    print(f"       Shift range: [{shift.min():.3f}, {shift.max():.3f}]")
    return mat_norm


# ==============================================================================
# 3. QC
# ==============================================================================

def plot_qc(mat: pd.DataFrame, design: pd.DataFrame, out_dir: str) -> list:
    """Génère les figures QC (détection, boxplots, RLE, valeurs manquantes)."""
    files = []
    conditions = design["condition"].values
    cmap = _condition_colors(conditions)

    # Fréquence de détection par peptide
    n_det = mat.notna().sum(axis=1)
    fig, ax = plt.subplots(figsize=(7, 4))
    ax.bar(range(len(n_det)), sorted(n_det.values, reverse=True),
           color="#3498DB", edgecolor="none")
    ax.set_xlabel("Peptides (sorted)"); ax.set_ylabel("Samples detected")
    ax.set_title("Detection frequency per peptide")
    fig.tight_layout()
    f = os.path.join(out_dir, "qc_frequency.png")
    fig.savefig(f, dpi=150); plt.close(fig); files.append(f)

    # Nb peptides détectés par échantillon
    n_per = mat.notna().sum(axis=0)
    fig, ax = plt.subplots(figsize=(max(8, len(n_per)*0.4), 4))
    ax.bar(range(len(n_per)), n_per.values,
           color=[cmap[c] for c in conditions], edgecolor="none")
    ax.set_xticks(range(len(n_per)))
    ax.set_xticklabels(design["label"].values, rotation=90, fontsize=7)
    ax.set_ylabel("Peptides detected"); ax.set_title("Detected peptides per sample")
    handles = [mpatches.Patch(color=v, label=k) for k, v in cmap.items()]
    ax.legend(handles=handles, fontsize=7)
    fig.tight_layout()
    f = os.path.join(out_dir, "qc_numbers.png")
    fig.savefig(f, dpi=150); plt.close(fig); files.append(f)

    # Boxplots (distribution log2 normalisée)
    fig, ax = plt.subplots(figsize=(max(10, len(mat.columns)*0.4), 5))
    ax.boxplot([mat[c].dropna().values for c in mat.columns],
               patch_artist=True,
               boxprops=dict(facecolor="#AED6F1"),
               medianprops=dict(color="red"),
               whiskerprops=dict(color="grey"),
               flierprops=dict(marker=".", markersize=2, alpha=0.3))
    ax.set_xticks(range(1, len(mat.columns)+1))
    ax.set_xticklabels(design["label"].values, rotation=90, fontsize=7)
    ax.set_ylabel("log2 Area (normalized)"); ax.set_title("Intensity distribution")
    fig.tight_layout()
    f = os.path.join(out_dir, "qc_normalization.png")
    fig.savefig(f, dpi=150); plt.close(fig); files.append(f)

    # RLE plot
    rle = mat.sub(mat.median(axis=1), axis=0)
    data_rle = [rle[c].dropna().values for c in mat.columns]
    med_rle = np.array([np.median(d) if len(d) else np.nan for d in data_rle])
    fig, ax = plt.subplots(figsize=(max(10, len(mat.columns)*0.4), 5))
    bp = ax.boxplot(data_rle, patch_artist=True, showfliers=False,
                    medianprops=dict(color="black", linewidth=1))
    for patch, c in zip(bp["boxes"], conditions):
        patch.set_facecolor(cmap[c]); patch.set_alpha(0.75)
    ax.axhline(0, color="red", ls="--", lw=0.8)
    ax.set_xticks(range(1, len(mat.columns)+1))
    ax.set_xticklabels(design["label"].values, rotation=90, fontsize=7)
    ax.set_ylabel("RLE (log2)")
    worst = float(np.nanmax(np.abs(med_rle))) if np.isfinite(med_rle).any() else 0
    ax.set_title(f"RLE plot (max |median| = {worst:.2f})")
    all_v = np.concatenate([d for d in data_rle if len(d)])
    if all_v.size:
        ylim = np.nanpercentile(np.abs(all_v), 99)
        ax.set_ylim(-ylim, ylim)
    handles = [mpatches.Patch(color=v, label=k) for k, v in cmap.items()]
    ax.legend(handles=handles, fontsize=7)
    fig.tight_layout()
    f = os.path.join(out_dir, "qc_rle.png")
    fig.savefig(f, dpi=150); plt.close(fig); files.append(f)

    # Missing values heatmap
    fig, ax = plt.subplots(figsize=(max(8, len(mat.columns)*0.35),
                                    min(12, len(mat)*0.02+2)))
    sns.heatmap(mat.isna().astype(int), cmap=["#2ECC71", "#E74C3C"],
                yticklabels=False, xticklabels=design["label"].values,
                ax=ax, cbar=False)
    ax.set_xticklabels(ax.get_xticklabels(), rotation=90, fontsize=7)
    ax.set_title("Missing values (red = absent)")
    fig.tight_layout()
    f = os.path.join(out_dir, "qc_missval.png")
    fig.savefig(f, dpi=150); plt.close(fig); files.append(f)

    return files


# ==============================================================================
# 4. FILTRAGE + IMPUTATION QRILC
# ==============================================================================

def filter_missval(mat: pd.DataFrame, design: pd.DataFrame, thr: int = 1
                   ) -> tuple[pd.DataFrame, np.ndarray]:
    """Filtre : peptide détecté dans >= (n_rep - thr) réplicats d'au moins 1 condition."""
    conditions = design["condition"].values
    keep = np.zeros(len(mat), dtype=bool)
    for cond in np.unique(conditions):
        cols = mat.columns[conditions == cond]
        n = len(cols)
        min_det = max(1, n - thr)
        keep |= (mat[cols].notna().sum(axis=1) >= min_det).values
    filtered = mat[keep].reset_index(drop=True)
    print(f"  -> Filter: {keep.sum()}/{len(mat)} peptides kept")
    return filtered, keep


def diagnose_missingness(mat: pd.DataFrame, design: pd.DataFrame,
                         out_dir: str) -> dict:
    """Diagnostic MNAR/MAR identique à deimos.py."""
    conditions = design["condition"].values
    arr = mat.values
    na = np.isnan(arr)
    n_missing = int(na.sum())
    if n_missing == 0:
        return {"n_missing": 0, "recommendation": "No imputation needed"}

    mnar, mar = 0, 0
    for i in range(arr.shape[0]):
        if not na[i].any():
            continue
        for cond in np.unique(conditions):
            cidx = np.where(conditions == cond)[0]
            sub = arr[i, cidx]
            n_miss = int(np.isnan(sub).sum())
            if n_miss == 0:
                continue
            if np.sum(~np.isnan(sub)) == 0:
                mnar += n_miss
            else:
                mar += n_miss

    total = mnar + mar
    pct_mnar = 100 * mnar / total if total else 0
    pct_mar  = 100 * mar  / total if total else 0

    mean_int = np.nanmean(arr, axis=1)
    prot_na = na.any(axis=1)
    delta = float(np.nanmedian(mean_int[~prot_na]) - np.nanmedian(mean_int[prot_na])) if prot_na.any() else 0

    fig, axes = plt.subplots(1, 2, figsize=(11, 4.2))
    axes[0].bar(["MNAR\n(full absence)", "MAR\n(partial)"],
                [pct_mnar, pct_mar], color=["#E8684A", "#5B8FF9"], edgecolor="black", lw=0.5)
    axes[0].set_ylabel("% missing values")
    axes[0].set_title(f"Missingness structure (n={total:,})", fontweight="bold")
    for x, v in enumerate([pct_mnar, pct_mar]):
        axes[0].text(x, v+1, f"{v:.0f}%", ha="center", fontweight="bold")
    axes[0].set_ylim(0, 105)

    bins = np.linspace(np.nanmin(mean_int), np.nanmax(mean_int), 40)
    axes[1].hist(mean_int[~prot_na], bins=bins, alpha=0.6,
                 color="#999", density=True, label="No missing")
    axes[1].hist(mean_int[prot_na], bins=bins, alpha=0.6,
                 color="#E8684A", density=True, label="With >= 1 missing")
    axes[1].set_xlabel("Mean log2 Area"); axes[1].set_ylabel("Density")
    axes[1].set_title("Intensity of peptides with gaps\n(left shift = MNAR)", fontweight="bold")
    axes[1].legend(fontsize=9)
    fig.tight_layout()
    f = os.path.join(out_dir, "diagnostic_missingness.png")
    fig.savefig(f, dpi=150, bbox_inches="tight"); plt.close(fig)

    print(f"  [STATS] Missingness: {pct_mnar:.0f}% MNAR / {pct_mar:.0f}% MAR "
          f"| delta intensity = {delta:.2f} log2")
    return {"n_missing": n_missing, "pct_mnar": round(pct_mnar,1),
            "pct_mar": round(pct_mar,1), "delta_intensity": round(delta,3),
            "plot": f}


def impute_qrilc(mat: pd.DataFrame, rng=None) -> pd.DataFrame:
    """QRILC — tirage dans la queue gauche tronquée de la distribution log2."""
    from scipy.stats import norm as _norm, truncnorm
    if rng is None:
        rng = np.random.default_rng()
    mat_imp = mat.copy()
    for col in mat_imp.columns:
        s = mat_imp[col]
        obs = s.dropna().values
        n_miss = s.isna().sum()
        if n_miss == 0 or len(obs) < 3:
            if n_miss > 0 and len(obs) > 0:
                mat_imp.loc[s.isna(), col] = obs.min()
            continue
        mu, sigma = np.mean(obs), np.std(obs)
        if sigma <= 0:
            mat_imp.loc[s.isna(), col] = mu
            continue
        p_miss = n_miss / len(s)
        q_censor = _norm.ppf(max(p_miss, 1e-4), loc=mu, scale=sigma)
        a, b = -np.inf, (q_censor - mu) / sigma
        draws = truncnorm.rvs(a, b, loc=mu, scale=sigma, size=n_miss,
                              random_state=rng)
        mat_imp.loc[s.isna(), col] = draws
    return mat_imp


def impute_minprob(mat: pd.DataFrame, q: float = 0.01, rng=None) -> pd.DataFrame:
    """MinProb — imputation bas-de-distribution (robustness iterations)."""
    draw = rng.normal if rng is not None else np.random.normal
    mat_imp = mat.copy()
    for col in mat_imp.columns:
        vals = mat_imp[col].dropna().values
        if len(vals) < 3:
            continue
        mu_low = np.quantile(vals, q)
        sd_low = np.std(vals) * 0.3
        n_miss = mat_imp[col].isna().sum()
        if n_miss > 0:
            mat_imp.loc[mat_imp[col].isna(), col] = draw(mu_low, sd_low, n_miss)
    return mat_imp


def plot_imputation(mat_filt: pd.DataFrame, mat_imp: pd.DataFrame,
                    out_dir: str) -> str:
    """Superposition densité avant/après imputation."""
    fig, ax = plt.subplots(figsize=(7, 4))
    v_before = mat_filt.values.flatten()
    v_before = v_before[~np.isnan(v_before)]
    v_after  = mat_imp.values.flatten()
    ax.hist(v_before, bins=80, alpha=0.6, color="#3498DB",
            label="Before imputation", density=True)
    ax.hist(v_after,  bins=80, alpha=0.6, color="#E74C3C",
            label="After imputation", density=True)
    ax.set_xlabel("log2 Area (normalized)"); ax.set_ylabel("Density")
    ax.set_title("QRILC imputation effect"); ax.legend()
    fig.tight_layout()
    f = os.path.join(out_dir, "qc_imputation.png")
    fig.savefig(f, dpi=150); plt.close(fig)
    return f


# ==============================================================================
# 5. ANALYSE DIFFÉRENTIELLE + ROBUSTNESS
# ==============================================================================

def run_differential_analysis(mat_imp: pd.DataFrame, mat_filt: pd.DataFrame,
                               meta: pd.DataFrame, design: pd.DataFrame,
                               params: dict, out_dir: str,
                               n_iter: int = 100
                               ) -> tuple[pd.DataFrame, list, list]:
    """
    limma eBayes sur les données peptidiques log2 normalisées et imputées.
    Pas de DEqMS (on est déjà au niveau peptide).
    Robustness score via re-imputations MinProb (même logique que deimos.py).
    """
    conditions = design["condition"].values
    expr = mat_imp.values.astype(float)   # (n_pep × n_samples)

    design_mat, group_names = make_design_matrix(conditions.tolist())
    contrast_mat, contrast_names = make_all_contrasts(group_names)

    print(f"\n[MODEL] {len(contrast_names)} contrasts | "
          f"{len(expr)} peptides | {expr.shape[1]} samples")

    fdr_global = params.get("fdr_global", False)
    fit   = lm_fit(expr, design_mat)
    fit_c = contrasts_fit(fit, contrast_mat)
    fit_e = ebayes(fit_c, fdr_global=fdr_global)
    print(f"  [OK] limma eBayes computed "
          f"(FDR {'global' if fdr_global else 'per contrast'})")

    # Moyennes par condition pour scatter plots
    cond_means = {}
    for cond in np.unique(conditions):
        idx = np.where(conditions == cond)[0]
        cond_clean = re.sub(r"[^A-Za-z0-9_]", ".", cond)
        cond_clean = "X" + cond_clean if cond_clean[0].isdigit() else cond_clean
        cond_means[cond_clean.replace(".", "_")] = expr[:, idx].mean(axis=1)

    df_results = meta.reset_index(drop=True).copy()

    scatter_files = []

    for i, cname in enumerate(contrast_names):
        tt = top_table(fit_e, i, protein_names=df_results["peptide_id"].values)
        df_results[f"{cname}_diff"]  = tt["logFC"].values
        df_results[f"{cname}_p.val"] = tt["P.Value"].values
        df_results[f"{cname}_p.adj"] = tt["adj.P.Val"].values
        p_for_pi = tt["P.Value"].values
        df_results[f"Pi_Score_{cname}"] = (
            np.abs(tt["logFC"].values) * (-np.log10(np.maximum(p_for_pi, 1e-10))))

    # --- Robustness score ---
    if n_iter and n_iter > 0:
        print(f"\n[...] Robustness ({n_iter} iterations)...")
        n_pep  = len(mat_filt)
        n_contr = len(contrast_names)
        p_key  = "adj.P.Val" if params["volcano_use_padj"] else "P.Value"
        lfc_min = params["volcano_lfc_min"]
        p_thr  = params["volcano_p_thresh"]

        def _one_iter(seed):
            rng = np.random.default_rng(seed)
            m_tmp = impute_minprob(mat_filt, rng=rng)
            f_tmp  = lm_fit(m_tmp.values.astype(float), design_mat)
            f_tmp_c = contrasts_fit(f_tmp, contrast_mat)
            f_tmp_e = ebayes(f_tmp_c, fdr_global=fdr_global)
            out = np.zeros((n_pep, n_contr), dtype=np.int8)
            for i in range(n_contr):
                tt = top_table(f_tmp_e, i)
                out[:, i] = ((np.abs(tt["logFC"].values) >= lfc_min) &
                             (tt[p_key].values < p_thr)).astype(np.int8)
            return out

        success = np.zeros((n_pep, n_contr), dtype=np.int32)
        try:
            from concurrent.futures import ThreadPoolExecutor
            n_workers = min(8, (os.cpu_count() or 2))
            done = 0
            with ThreadPoolExecutor(max_workers=n_workers) as ex:
                for res in ex.map(_one_iter, range(n_iter)):
                    success += res
                    done += 1
                    if done % 10 == 0 or done == n_iter:
                        print(f"    {done}/{n_iter}", end="\r")
            print()
        except Exception as e:
            print(f"  (sequential: {e})")
            for k in range(n_iter):
                success += _one_iter(k)
        for i, cname in enumerate(contrast_names):
            df_results[f"Robustness_Score_{cname}"] = success[:, i]
        print("  [OK] Robustness done")
    else:
        print("\n[SKIP] Robustness disabled")

    # --- Scatter plots ---
    for i, cname in enumerate(contrast_names):
        try:
            g1, g2 = cname.split("_vs_")
            if g1 in cond_means and g2 in cond_means:
                xv, yv = cond_means[g1], cond_means[g2]
                lfc = df_results[f"{cname}_diff"].values
                pv  = df_results[f"{cname}_p.val"].values
                status = np.full(len(lfc), "Not changed", dtype=object)
                status[(lfc >  params["volcano_lfc_min"]) & (pv < 0.05)] = "Up regulated"
                status[(lfc < -params["volcano_lfc_min"]) & (pv < 0.05)] = "Down regulated"
                col_map = {"Up regulated": "#f3a583",
                           "Down regulated": "#92dadd",
                           "Not changed": "#bdbdbd"}
                fig, ax = plt.subplots(figsize=(6, 6))
                for s, col in col_map.items():
                    m = status == s
                    ax.scatter(xv[m], yv[m], c=col, s=10, alpha=0.5, label=s)
                lims = [min(xv.min(), yv.min()), max(xv.max(), yv.max())]
                ax.plot(lims, lims, "k--", lw=0.8, alpha=0.5)
                ax.set_xlabel(f"log2 Area — {g1}")
                ax.set_ylabel(f"log2 Area — {g2}")
                ax.set_title(f"Scatter: {g1} vs {g2}")
                ax.legend(fontsize=8)
                fig.tight_layout()
                f = os.path.join(out_dir, f"scatter_{cname}.png")
                fig.savefig(f, dpi=150); plt.close(fig)
                scatter_files.append(f)
        except Exception as e:
            print(f"  [WARN] Scatter {cname}: {e}")

    return df_results, contrast_names, scatter_files


# ==============================================================================
# 6. VOLCANO PLOTS
# ==============================================================================

def plot_volcanoes(df_results: pd.DataFrame, contrast_names: list,
                   params: dict, out_dir: str) -> tuple[list, str]:
    """Volcanos individuels + facet — identique à deimos.py, adapté peptide."""
    p_key  = "p.adj" if params["volcano_use_padj"] else "p.val"
    thresh = params["volcano_p_thresh"]
    lfc    = params["volcano_lfc_min"]
    col_map = {"Up regulated": "#E74C3C",
               "Down regulated": "#3498DB",
               "Not significant": "#BDBDBD"}
    indiv_files = []

    for cname in contrast_names:
        diff_col = f"{cname}_diff"
        pval_col = f"{cname}_{p_key}"
        if diff_col not in df_results.columns:
            continue
        d  = df_results[diff_col].values
        p  = df_results[pval_col].values
        nm = df_results["peptide_id"].values
        y  = -np.log10(np.maximum(p, 1e-10))

        status = np.full(len(d), "Not significant", dtype=object)
        status[(d >  lfc) & (p < thresh)] = "Up regulated"
        status[(d < -lfc) & (p < thresh)] = "Down regulated"

        fig, ax = plt.subplots(figsize=(7, 6))
        for s, col in col_map.items():
            m = status == s
            ax.scatter(d[m], y[m], c=col, s=10, alpha=0.7, label=s)

        # Labels des significatifs (top 40 par score)
        sig = status != "Not significant"
        xs, ys, ns = d[sig], y[sig], nm[sig]
        if len(ns) > 40:
            score = np.abs(xs) + ys
            keep = np.argsort(score)[::-1][:40]
            xs, ys, ns = xs[keep], ys[keep], ns[keep]
        # Afficher seulement la séquence (avant le premier |)
        labels_txt = [n.split("|")[0] for n in ns]
        texts = [ax.text(xi, yi, li, fontsize=5, alpha=0.85)
                 for xi, yi, li in zip(xs, ys, labels_txt)]
        try:
            from adjustText import adjust_text
            adjust_text(texts, ax=ax,
                        arrowprops=dict(arrowstyle="-", color="grey",
                                        lw=0.4, shrinkA=4, shrinkB=2))
        except ImportError:
            pass

        ax.axvline(-lfc, ls="--", lw=0.8, color="grey")
        ax.axvline( lfc, ls="--", lw=0.8, color="grey")
        ax.axhline(-np.log10(thresh), ls="--", lw=0.8, color="grey")
        ax.set_xlabel("log2 Fold Change")
        ax.set_ylabel(f"-log10({'p.adj' if params['volcano_use_padj'] else 'p.value'})")
        ax.set_title(cname.replace("_vs_", " vs "))
        ax.legend(fontsize=7)
        fig.tight_layout()
        f = os.path.join(out_dir, f"volc_{cname}.png")
        fig.savefig(f, dpi=150); plt.close(fig)
        indiv_files.append(f)

    # Facet volcano
    n_c, n_cols = len(contrast_names), 4
    n_rows = (n_c + n_cols - 1) // n_cols
    fig, axes = plt.subplots(n_rows, n_cols,
                              figsize=(n_cols*4, n_rows*3.5))
    axes = np.array(axes).flatten()
    for idx, cname in enumerate(contrast_names):
        ax = axes[idx]
        if f"{cname}_diff" not in df_results.columns:
            ax.set_visible(False); continue
        d = df_results[f"{cname}_diff"].values
        p = df_results[f"{cname}_{p_key}"].values
        p_raw = df_results[f"{cname}_p.val"].values
        status = np.full(len(d), "NS", dtype=object)
        status[(d >  lfc) & (p < thresh)] = "Up"
        status[(d < -lfc) & (p < thresh)] = "Down"
        colors_pt = [{"Up": "#E74C3C", "Down": "#3498DB", "NS": "#DDD"}[s] for s in status]
        ax.scatter(d, -np.log10(np.maximum(p_raw, 1e-10)),
                   c=colors_pt, s=4, alpha=0.6)
        ax.axvline(-lfc, ls=":", lw=0.6, color="black", alpha=0.4)
        ax.axvline( lfc, ls=":", lw=0.6, color="black", alpha=0.4)
        ax.axhline(-np.log10(thresh), ls=":", lw=0.6, color="black", alpha=0.4)
        ax.set_title(cname.replace("_vs_", " vs "), fontsize=7, fontweight="bold",
                     color="white", bbox=dict(facecolor="#34495E", boxstyle="round,pad=0.2"))
        ax.set_xlabel("log2FC", fontsize=7); ax.set_ylabel("-log10(p)", fontsize=7)
        ax.tick_params(labelsize=6)
    for idx in range(n_c, len(axes)):
        axes[idx].set_visible(False)
    patches = [mpatches.Patch(color="#E74C3C", label="Up"),
               mpatches.Patch(color="#3498DB", label="Down"),
               mpatches.Patch(color="#DDD", label="NS")]
    fig.legend(handles=patches, loc="lower right", fontsize=9, ncol=3)
    fig.suptitle(f"Facet Volcano — {n_c} contrasts | "
                 f"{'p.adj' if params['volcano_use_padj'] else 'p.val'} < {thresh} "
                 f"| |log2FC| > {lfc:.2f}",
                 fontsize=11, fontweight="bold")
    fig.tight_layout(rect=[0, 0.02, 1, 0.97])
    f_facet = os.path.join(out_dir, "all_volcano_plots_grid.png")
    fig.savefig(f_facet, dpi=150, bbox_inches="tight"); plt.close(fig)
    print(f"  [OK] Facet volcano saved ({n_c} contrasts)")
    return indiv_files, f_facet


# ==============================================================================
# 7. PCA + UMAP + CORRÉLATION
# ==============================================================================

def plot_pca(mat_imp: pd.DataFrame, design: pd.DataFrame, out_dir: str) -> list:
    expr = mat_imp.T.values
    imp  = SimpleImputer(strategy="mean")
    expr_imp = StandardScaler().fit_transform(imp.fit_transform(expr))
    pca = PCA(n_components=min(10, *expr_imp.shape))
    coords = pca.fit_transform(expr_imp)
    var_exp = pca.explained_variance_ratio_
    conditions = design["condition"].values
    labels = design["label"].values
    cmap = _condition_colors(conditions)
    files = []
    for with_ell in [False, True]:
        fig, ax = plt.subplots(figsize=(9, 7))
        for cond in np.unique(conditions):
            m = conditions == cond
            ax.scatter(coords[m,0], coords[m,1], color=cmap[cond],
                       label=cond, s=60, alpha=0.85, zorder=3)
            for xi, yi, li in zip(coords[m,0], coords[m,1], labels[m]):
                ax.annotate(li, (xi,yi), fontsize=6, alpha=0.7,
                            xytext=(3,3), textcoords="offset points")
            if with_ell and m.sum() >= 3:
                _draw_ellipse(ax, coords[m,0], coords[m,1], color=cmap[cond])
        ax.set_xlabel(f"PC1 ({var_exp[0]*100:.1f}%)")
        ax.set_ylabel(f"PC2 ({var_exp[1]*100:.1f}%)")
        ax.set_title("PCA — peptide-level (imputed data)"
                     + (" — 95% ellipses" if with_ell else ""))
        ax.legend(fontsize=7, bbox_to_anchor=(1.01,1), loc="upper left")
        ax.axhline(0, color="grey", lw=0.5); ax.axvline(0, color="grey", lw=0.5)
        fig.tight_layout()
        f = os.path.join(out_dir, f"plot_pca{'_ellipses' if with_ell else ''}.png")
        fig.savefig(f, dpi=150, bbox_inches="tight"); plt.close(fig)
        files.append(f)
    return files


def _draw_ellipse(ax, x, y, color, alpha=0.15, level=0.95):
    from matplotlib.patches import Ellipse
    from scipy.stats import chi2
    cov = np.cov(x, y)
    vals, vecs = np.linalg.eigh(cov)
    order = vals.argsort()[::-1]; vals, vecs = vals[order], vecs[:, order]
    angle = np.degrees(np.arctan2(*vecs[:,0][::-1]))
    chi2_v = chi2.ppf(level, df=2)
    w, h = 2*np.sqrt(vals*chi2_v)
    ell = mpatches.Ellipse(xy=(x.mean(), y.mean()),
                            width=w, height=h, angle=angle,
                            color=color, alpha=alpha)
    ax.add_patch(ell)


def plot_correlation_heatmap(mat_imp: pd.DataFrame, design: pd.DataFrame,
                              out_dir: str) -> str:
    corr = np.corrcoef(mat_imp.values.T)
    labels = design["label"].values
    conditions = design["condition"].values
    cmap = _condition_colors(conditions)
    fig, ax = plt.subplots(figsize=(max(8, len(labels)*0.45),
                                     max(7, len(labels)*0.4)))
    im = ax.imshow(corr, cmap="RdBu_r", vmin=corr.min()*0.95, vmax=1.0)
    for i in range(len(labels)):
        for j in range(len(labels)):
            ax.text(j, i, f"{corr[i,j]:.2f}", ha="center", va="center",
                    fontsize=max(5, 8-len(labels)//5))
    ax.set_xticks(range(len(labels))); ax.set_yticks(range(len(labels)))
    ax.set_xticklabels(labels, rotation=90, fontsize=7)
    ax.set_yticklabels(labels, fontsize=7)
    for j, cond in enumerate(conditions):
        ax.add_patch(mpatches.Rectangle((j-.5,-1.5),1,.8,color=cmap[cond],clip_on=False))
    plt.colorbar(im, ax=ax, fraction=0.03)
    ax.set_title("Pearson correlation — peptide-level")
    fig.tight_layout()
    f = os.path.join(out_dir, "heatmap_correlation.png")
    fig.savefig(f, dpi=150, bbox_inches="tight"); plt.close(fig)
    return f


def plot_umap(mat_imp: pd.DataFrame, design: pd.DataFrame,
              out_dir: str) -> tuple[str, pd.DataFrame]:
    expr = mat_imp.T.values
    imp  = SimpleImputer(strategy="mean")
    expr_imp = imp.fit_transform(expr)
    n_nb = min(5, expr_imp.shape[0]-1)
    emb  = umap.UMAP(n_neighbors=n_nb, min_dist=0.1, random_state=42).fit_transform(expr_imp)
    conditions = design["condition"].values
    labels = design["label"].values
    cmap = _condition_colors(conditions)
    fig, ax = plt.subplots(figsize=(11, 7))
    for cond in np.unique(conditions):
        m = conditions == cond
        ax.scatter(emb[m,0], emb[m,1], color=cmap[cond], label=cond, s=60, alpha=0.85)
        for xi, yi, li in zip(emb[m,0], emb[m,1], labels[m]):
            ax.annotate(li, (xi,yi), fontsize=6, alpha=0.7,
                        xytext=(3,3), textcoords="offset points")
    ax.set_xlabel("UMAP1"); ax.set_ylabel("UMAP2")
    ax.set_title(f"Peptide UMAP ({mat_imp.shape[0]} peptides)")
    ax.legend(fontsize=7, bbox_to_anchor=(1.01,1), loc="upper left")
    fig.tight_layout()
    f = os.path.join(out_dir, "umap_plot.png")
    fig.savefig(f, dpi=150, bbox_inches="tight"); plt.close(fig)
    print("  [OK] UMAP generated")
    umap_coords = pd.DataFrame({"label": labels, "condition": conditions,
                                 "UMAP1": emb[:,0], "UMAP2": emb[:,1]})
    return f, umap_coords


# ==============================================================================
# 8. ANOVA + HEATMAPS (niveau peptide)
# ==============================================================================

def run_anova_heatmaps(mat_imp: pd.DataFrame, meta: pd.DataFrame,
                       design: pd.DataFrame, params: dict,
                       out_dir: str) -> tuple:
    """ANOVA one-way peptide × condition + heatmaps Z-score."""
    from scipy.stats import f as fdist

    conditions = design["condition"].values
    groups = np.unique(conditions)
    expr = mat_imp.values.astype(float)
    n_pep, n_samp = expr.shape
    n_groups = len(groups)
    df_between = n_groups - 1
    df_within  = n_samp - n_groups

    grand_mean = expr.mean(axis=1, keepdims=True)
    ss_between = np.zeros(n_pep)
    ss_within  = np.zeros(n_pep)
    for g in groups:
        idx = np.where(conditions == g)[0]
        gm = expr[:, idx].mean(axis=1, keepdims=True)
        ss_between += len(idx) * (gm.flatten() - grand_mean.flatten())**2
        ss_within  += ((expr[:, idx] - gm)**2).sum(axis=1)

    ms_between = ss_between / df_between
    ms_within  = np.maximum(ss_within / df_within, 1e-12)
    f_stat = ms_between / ms_within
    p_vals = 1 - fdist.cdf(f_stat, df_between, df_within)
    p_adj  = _bh_correction(p_vals)

    df_anova = meta.reset_index(drop=True).copy()
    df_anova["F.stat"]  = f_stat
    df_anova["p.value"] = p_vals
    df_anova["p.adj"]   = p_adj
    p_key_col = "p.adj" if params["anova_use_padj"] else "p.value"
    df_anova["significant"] = df_anova[p_key_col] < params["anova_p_thresh"]
    df_anova = df_anova.sort_values("p.value").reset_index(drop=True)

    sig_ids = df_anova.loc[df_anova["significant"], "peptide_id"].values
    print(f"  -> ANOVA: {len(sig_ids)} significant peptides "
          f"({p_key_col} < {params['anova_p_thresh']})")

    hm_classic = hm_clusters = hm_violin = None
    cluster_mapping = []
    mat_zscore = None

    if len(sig_ids) > 0:
        sig_idx = meta.index[meta["peptide_id"].isin(sig_ids)].tolist()
        mat_sig = expr[sig_idx, :]
        mat_z   = (mat_sig - mat_sig.mean(axis=1, keepdims=True)) / np.maximum(
                   mat_sig.std(axis=1, keepdims=True), 1e-8)
        mat_zscore = mat_z

        sample_labels = design["label"].values
        cmap_cond = _condition_colors(conditions)

        # Heatmap classique
        hm_classic = _draw_heatmap_pch(
            mat_z, sig_ids, sample_labels, conditions, cmap_cond,
            title=f"ANOVA peptides (n={len(sig_ids)}, {p_key_col} < {params['anova_p_thresh']})",
            out_path=os.path.join(out_dir, "heatmap_annotated.png"),
            show_row_names=len(sig_ids) <= 80
        )

        # Heatmap clusters
        n_req = params.get("n_heatmap_clusters", 3)
        n_row_k = max(2, min(n_req, len(sig_ids) // 3)) if len(sig_ids) >= 6 else 1
        n_col_k = min(4, len(groups))
        if n_row_k < n_req:
            print(f"  [WARN] {len(sig_ids)} sig peptides -> {n_row_k} clusters")
        _, _, row_labels = _cluster_heatmap(mat_z, n_row_k, n_col_k)

        hm_clusters = _draw_heatmap_pch(
            mat_z, sig_ids, sample_labels, conditions, cmap_cond,
            title=f"Clustered heatmap (n={len(sig_ids)})",
            out_path=os.path.join(out_dir, "heatmap_clusters.png"),
            show_row_names=len(sig_ids) <= 80,
            row_split_labels=row_labels
        )

        for ci in np.unique(row_labels):
            pep_c = sig_ids[row_labels == ci]
            cluster_mapping.extend([(p, f"Cluster_{ci+1}") for p in pep_c])
        df_anova["Cluster_ID"] = df_anova["peptide_id"].map(dict(cluster_mapping)).fillna("")

        hm_violin = _plot_cluster_violin(
            mat_z, sig_ids, row_labels, conditions, design,
            out_path=os.path.join(out_dir, "plt_clusters_profiles.png")
        )

    return df_anova, hm_classic, hm_clusters, hm_violin, mat_zscore, cluster_mapping, sig_ids


# Réutilisation directe des fonctions heatmap de deimos (même signature)
def _draw_heatmap_pch(mat_z, row_names, sample_labels, conditions, cmap_cond,
                      title, out_path, show_row_names=True, row_split_labels=None):
    try:
        import PyComplexHeatmap as pch
    except ImportError:
        return _draw_heatmap(mat_z, row_names, sample_labels, conditions,
                             cmap_cond, title, out_path,
                             show_row_names=show_row_names,
                             row_cluster_labels=row_split_labels)

    n_prot, n_samp = mat_z.shape
    df = pd.DataFrame(mat_z, index=list(row_names), columns=list(sample_labels))
    cond_series = pd.Series(list(conditions), index=list(sample_labels))
    col_anno = pch.HeatmapAnnotation(
        Condition=pch.anno_simple(cond_series, colors=cmap_cond, height=4,
                                   legend=True, add_text=False),
        axis=1, verbose=0, label_side="right",
        label_kws={"fontsize": 9, "fontweight": "bold",
                   "horizontalalignment": "left"})
    row_split = None; left_anno = None
    if row_split_labels is not None:
        import collections as _c
        counts = _c.Counter(row_split_labels)
        if len(row_split_labels) >= 6 and all(v >= 2 for v in counts.values()):
            cn = [f"C{l+1}" for l in row_split_labels]
            row_split = pd.Series(cn, index=list(row_names))
            uniq = sorted(set(cn))
            pal = dict(zip(uniq, ["#5B8FF9","#F6BD16","#5AD8A6","#E8684A",
                                   "#9270CA","#FF9D4D","#269A99","#FF99C3"][:len(uniq)]))
            left_anno = pch.HeatmapAnnotation(
                Cluster=pch.anno_simple(row_split, colors=pal, height=4,
                                         legend=True, add_text=True,
                                         text_kws={"fontsize": 8}),
                axis=0, verbose=0, label_side="top",
                label_kws={"fontsize": 9, "fontweight": "bold"})

    fig_h = max(6, min(n_prot * 0.18 + 3, 28))
    fig_w = max(9, n_samp * 0.40 + 5)
    plt.figure(figsize=(fig_w, fig_h))
    try:
        pch.ClusterMapPlotter(data=df, top_annotation=col_anno,
                               left_annotation=left_anno,
                               row_cluster=True, col_cluster=True,
                               row_split=row_split,
                               row_split_gap=2.2, col_split_gap=1.2,
                               row_dendrogram=True, col_dendrogram=True,
                               show_rownames=show_row_names, show_colnames=True,
                               row_names_side="right", cmap="RdBu_r",
                               vmin=-2.5, vmax=2.5, center=0,
                               label="Z-score", legend=True,
                               xticklabels_kws={"labelrotation": 90, "labelsize": 7},
                               yticklabels_kws={"labelsize": 6},
                               verbose=0)
        plt.suptitle(title, fontsize=12, fontweight="bold", y=1.01)
        plt.savefig(out_path, dpi=160, bbox_inches="tight", facecolor="white")
        plt.close("all")
        return out_path
    except Exception as e:
        plt.close("all")
        print(f"  [WARN] PyComplexHeatmap failed ({e}) -> fallback")
        return _draw_heatmap(mat_z, row_names, sample_labels, conditions,
                             cmap_cond, title, out_path,
                             show_row_names=show_row_names,
                             row_cluster_labels=row_split_labels)


def _draw_heatmap(mat_z, row_names, sample_labels, conditions, cmap_cond,
                  title, out_path, show_row_names=True, row_cluster_labels=None):
    n_prot, n_samp = mat_z.shape
    fig = plt.figure(figsize=(max(8, n_samp*0.35+3), max(6, min(n_prot*0.18+2, 25))))
    gs  = fig.add_gridspec(2, 2, height_ratios=[0.04,1],
                            width_ratios=[1,0.05], hspace=0.01)
    ax_ann  = fig.add_subplot(gs[0,0])
    ax_heat = fig.add_subplot(gs[1,0])
    ax_cbar = fig.add_subplot(gs[1,1])
    cond_colors = [cmap_cond[c] for c in conditions]
    for j, col in enumerate(cond_colors):
        ax_ann.add_patch(mpatches.Rectangle((j,0),1,1,color=col))
    ax_ann.set_xlim(0,n_samp); ax_ann.set_ylim(0,1); ax_ann.axis("off")
    ax_ann.set_title(title, fontsize=9, pad=22)
    # Légende des groupes (un patch par condition, ordre d'apparition)
    seen = list(dict.fromkeys(conditions))
    legend_handles = [mpatches.Patch(color=cmap_cond[c], label=str(c))
                      for c in seen]
    ax_ann.legend(handles=legend_handles, loc="lower left",
                  bbox_to_anchor=(0, 1.15), ncol=min(len(seen), 6),
                  fontsize=7, frameon=False, handlelength=1.2,
                  handleheight=1.2, columnspacing=1.0, title="Group",
                  title_fontsize=7)
    im = ax_heat.imshow(mat_z, aspect="auto", cmap="RdBu_r",
                        vmin=-2.5, vmax=2.5, interpolation="nearest")
    ax_heat.set_xticks(range(n_samp))
    ax_heat.set_xticklabels(sample_labels, rotation=90, fontsize=6)
    if show_row_names:
        ax_heat.set_yticks(range(n_prot))
        ax_heat.set_yticklabels(row_names, fontsize=5)
    else:
        ax_heat.set_yticks([])
    if row_cluster_labels is not None:
        prev = row_cluster_labels[0]
        for k, cl in enumerate(row_cluster_labels):
            if cl != prev:
                ax_heat.axhline(k-0.5, color="white", lw=1.5)
                prev = cl
    plt.colorbar(im, cax=ax_cbar, label="Z-score")
    fig.tight_layout()
    fig.savefig(out_path, dpi=150, bbox_inches="tight"); plt.close(fig)
    return out_path


def _cluster_heatmap(mat_z, n_row_k, n_col_k):
    from sklearn.cluster import KMeans
    km_r = KMeans(n_clusters=n_row_k, random_state=42, n_init=10)
    row_labels = km_r.fit_predict(mat_z)
    km_c = KMeans(n_clusters=n_col_k, random_state=42, n_init=10)
    col_labels = km_c.fit_predict(mat_z.T)
    return np.argsort(row_labels), np.argsort(col_labels), row_labels


def _plot_cluster_violin(mat_z, row_names, row_labels, conditions, design, out_path):
    n_clusters = len(np.unique(row_labels))
    fig, axes = plt.subplots(1, n_clusters, figsize=(n_clusters*4, 5), sharey=False)
    if n_clusters == 1:
        axes = [axes]
    cmap = _condition_colors(conditions)
    for ci, ax in enumerate(axes):
        m = row_labels == ci
        data_cl = mat_z[m, :]
        df_v = pd.DataFrame(data_cl.T, columns=row_names[m])
        df_v["condition"] = conditions
        df_long = df_v.melt(id_vars=["condition"], var_name="peptide", value_name="Z")
        for k, cond in enumerate(sorted(np.unique(conditions))):
            sub = df_long[df_long["condition"] == cond]["Z"].values
            if len(sub) > 1:
                parts = ax.violinplot([sub], positions=[k], widths=0.6, showmedians=False)
                for pc in parts["bodies"]:
                    pc.set_facecolor(cmap[cond]); pc.set_alpha(0.6)
            q1, med, q3 = np.percentile(sub, [25,50,75])
            ax.plot([k-.1,k+.1], [med,med], "k-", lw=2)
            ax.plot([k,k], [q1,q3], "k-", lw=1.5)
        ax.set_xticks(range(len(np.unique(conditions))))
        ax.set_xticklabels(sorted(np.unique(conditions)), rotation=45, ha="right", fontsize=7)
        ax.set_title(f"Cluster {ci+1} (n={m.sum()} peptides)", fontsize=9)
        ax.set_ylabel("Z-score" if ci == 0 else "")
    fig.suptitle("Peptide expression profiles by cluster", fontsize=11)
    fig.tight_layout()
    fig.savefig(out_path, dpi=150, bbox_inches="tight"); plt.close(fig)
    return out_path


# ==============================================================================
# 9. UPSET PLOT
# ==============================================================================

def plot_upset(df_results: pd.DataFrame, contrast_names: list,
               params: dict, out_dir: str) -> tuple[str, pd.DataFrame]:
    """UpSet des peptides DEP par contraste."""
    p_key  = "p.adj" if params["volcano_use_padj"] else "p.val"
    thresh = params["volcano_p_thresh"]
    lfc    = params["volcano_lfc_min"]

    upset_dict = {}
    for cname in contrast_names:
        d_col = f"{cname}_diff"
        p_col = f"{cname}_{p_key}"
        if d_col not in df_results.columns:
            continue
        mask = ((np.abs(df_results[d_col].values) >= lfc) &
                (df_results[p_col].values < thresh))
        prots = df_results.loc[mask, "peptide_id"].dropna().unique().tolist()
        if prots:
            upset_dict[cname] = set(prots)

    if len(upset_dict) < 2:
        print("  [WARN] Not enough DEP contrasts for UpSet")
        return None, pd.DataFrame()

    all_p = sorted(set.union(*upset_dict.values()))
    cnames_list = list(upset_dict.keys())
    n_sets = len(cnames_list)

    mat_bin = pd.DataFrame(
        {k: [1 if p in v else 0 for p in all_p] for k, v in upset_dict.items()},
        index=all_p)

    inter_counts = {}
    for _, row in mat_bin.iterrows():
        key = tuple(cnames_list[j] for j in range(n_sets) if row.iloc[j] == 1)
        if key:
            inter_counts[key] = inter_counts.get(key, 0) + 1
    top_inter = sorted(inter_counts.items(), key=lambda x: -x[1])[:30]
    n_inter = len(top_inter)

    fig = plt.figure(figsize=(max(14, n_inter*0.5), 8))
    gs  = fig.add_gridspec(2, 2, width_ratios=[1,4], height_ratios=[2,1], hspace=0.05)
    ax_bar  = fig.add_subplot(gs[0,1])
    ax_mat  = fig.add_subplot(gs[1,1])
    ax_sets = fig.add_subplot(gs[1,0])

    ax_bar.bar(range(n_inter), [v for _,v in top_inter], color="#2C3E50")
    ax_bar.set_ylabel("DEP peptides"); ax_bar.set_xticks([])
    ax_bar.set_title("UpSet Plot — DEP peptide intersections")

    ax_mat.set_xlim(-0.5, n_inter-0.5); ax_mat.set_ylim(-0.5, n_sets-0.5)
    ax_mat.set_yticks(range(n_sets))
    ax_mat.set_yticklabels([c[:25] for c in cnames_list], fontsize=6)
    ax_mat.set_xticks([])
    for xi, (keys, _) in enumerate(top_inter):
        active = [cnames_list.index(k) for k in keys if k in cnames_list]
        for yi in active:
            ax_mat.scatter(xi, yi, s=50, c="#2C3E50", zorder=3)
        if len(active) > 1:
            ax_mat.plot([xi,xi], [min(active),max(active)], color="#2C3E50", lw=2)

    ax_sets.barh(range(n_sets), [len(upset_dict[c]) for c in cnames_list], color="#3498DB")
    ax_sets.invert_xaxis(); ax_sets.set_yticks([])
    ax_sets.set_xlabel("Total DEP")
    fig.tight_layout()
    f_upset = os.path.join(out_dir, "upset_plot.png")
    fig.savefig(f_upset, dpi=150, bbox_inches="tight"); plt.close(fig)

    mat_bin["Nb_Conditions"] = mat_bin.sum(axis=1)
    mat_bin_X = mat_bin.copy()
    for col in cnames_list:
        mat_bin_X[col] = mat_bin_X[col].map({1:"X", 0:""})
    mat_bin_X = (mat_bin_X.sort_values("Nb_Conditions", ascending=False)
                           .reset_index().rename(columns={"index": "peptide_id"}))
    print(f"  [OK] UpSet done ({len(upset_dict)} contrasts)")
    return f_upset, mat_bin_X


# ==============================================================================
# 10. AGRÉGATION PROTÉINE (optionnelle, post-hoc)
# ==============================================================================

def aggregate_to_protein(df_results: pd.DataFrame, mat_imp: pd.DataFrame,
                          design: pd.DataFrame, contrast_names: list,
                          params: dict, out_dir: str) -> pd.DataFrame:
    """
    Agrégation median-polish des intensités peptidiques → niveau protéine.
    Utile pour les visualisations globales (heatmap protéine, boxplot rapide).
    Ne remplace PAS l'analyse différentielle peptide-level.
    Retourne un DataFrame protéine × sample (log2, normalisé).
    """
    if "Accession" not in df_results.columns:
        print("  [WARN] No 'Accession' column — protein aggregation skipped")
        return None

    print("\n[AGG] Aggregating peptides to protein level (median)...")
    df = df_results[["peptide_id", "Accession"]].copy()
    df["Accession"] = df["Accession"].fillna("Unknown").astype(str)

    mat_with_id = mat_imp.copy()
    mat_with_id.index = df["peptide_id"].values
    mat_with_id["_acc"] = df["Accession"].values

    # Médiane par protéine × échantillon
    prot_mat = (mat_with_id.groupby("_acc")
                .median(numeric_only=True)
                .reset_index()
                .rename(columns={"_acc": "Accession"}))

    # Distribution très asymétrique (1 protéine peut avoir des centaines de
    # peptides). On borne l'axe X sur l'essentiel (P99) pour rester lisible et
    # on regroupe la queue dans une barre "overflow". Bins de largeur 1 sur la
    # plage informative.
    fig, ax = plt.subplots(figsize=(max(8, len(mat_imp.columns)*0.4), 4))
    n_pep_per_prot = df.groupby("Accession")["peptide_id"].count()
    vals = n_pep_per_prot.values
    p_cap = int(np.percentile(vals, 95))
    x_cap = max(p_cap, 10)                      # au moins 0–10 visible
    n_over = int((vals > x_cap).sum())        # protéines au-delà du cap
    clipped = np.clip(vals, 0, x_cap + 1)     # la queue tombe dans le dernier bin
    bins = np.arange(0.5, x_cap + 2.5, 1)     # 1 bin par valeur entière
    ax.hist(clipped, bins=bins, color="#3498DB", edgecolor="white")
    ax.set_xlabel("Peptides per protein")
    ax.set_ylabel("Count")
    ax.set_xlim(0.5, x_cap + 1.5)
    # Marquer la barre d'overflow si des protéines dépassent le cap
    if n_over > 0:
        ax.axvline(x_cap + 0.5, color="#E74C3C", ls="--", lw=0.8)
        ax.text(x_cap + 0.5, ax.get_ylim()[1]*0.96,
                f"  >{x_cap}: {n_over} prot. (max={vals.max()})",
                fontsize=7, color="#E74C3C", ha="right", va="top")
    ax.set_title(f"Peptide aggregation ({len(prot_mat)} proteins, "
                 f"median={n_pep_per_prot.median():.1f}, "
                 f"max={vals.max()} pep/prot)")
    fig.tight_layout()
    f = os.path.join(out_dir, "protein_aggregation_summary.png")
    fig.savefig(f, dpi=150); plt.close(fig)

    print(f"  -> {len(prot_mat)} proteins from {len(df)} peptides "
          f"(median {n_pep_per_prot.median():.1f} pep/prot)")
    return prot_mat


# ==============================================================================
# 11. EXPORT EXCEL
# ==============================================================================

def build_methods_sheet(ws, params: dict, n_raw_pep: int, n_filt_pep: int,
                         n_sig_anova: int, contrast_names: list,
                         design: pd.DataFrame, n_iter: int):
    """Feuille Méthodes adaptée au pipeline peptide DDA/Peaks."""
    from openpyxl.styles import Alignment, Border, Side

    tf   = Font(bold=True, size=14, color="FFFFFF")
    tf2  = PatternFill("solid", fgColor="2C3E50")
    sf   = Font(bold=True, size=11, color="FFFFFF")
    sf2  = PatternFill("solid", fgColor="34495E")
    kf   = Font(bold=True, size=10)
    kf2  = PatternFill("solid", fgColor="ECF0F1")
    wrap = Alignment(wrap_text=True, vertical="top")
    thin = Side(style="thin", color="BDC3C7")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 70

    row = 1
    def t(text):
        nonlocal row
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        c = ws.cell(row, 1, text); c.font = tf; c.fill = tf2
        c.alignment = Alignment(vertical="center"); ws.row_dimensions[row].height = 24
        row += 2

    def s(text):
        nonlocal row
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        c = ws.cell(row, 1, text); c.font = sf; c.fill = sf2
        ws.row_dimensions[row].height = 18; row += 1
        for col, h in enumerate(["Item","Choice / Value","Method / Detail"],1):
            ws.cell(row, col, h).font = Font(bold=True, italic=True, size=9, color="7F8C8D")
        row += 1

    def kv(k, v, m=""):
        nonlocal row
        c1 = ws.cell(row,1,k); c1.font=kf; c1.fill=kf2; c1.alignment=wrap; c1.border=bdr
        c2 = ws.cell(row,2,v); c2.alignment=wrap; c2.border=bdr
        c3 = ws.cell(row,3,m); c3.alignment=wrap; c3.border=bdr
        row += 1

    p_volc  = "p.adj (FDR-BH)" if params["volcano_use_padj"] else "p.value (raw)"
    p_anova = "p.adj (FDR-BH)" if params["anova_use_padj"] else "p.value (raw)"
    n_samp = len(design); n_cond = design["condition"].nunique()
    reps = design.groupby("condition").size()
    rep_r = f"{reps.min()}–{reps.max()}"

    t("PHOBOS — DDA Peptide Pipeline — Methods & Parameters")
    s("1. EXPERIMENTAL DESIGN")
    kv("Samples", str(n_samp), "Area columns from Peaks protein-peptides.csv")
    kv("Conditions", str(n_cond), ", ".join(map(str, design["condition"].unique())))
    kv("Replicates", rep_r, "Biological replicates per condition")
    kv("Comparisons", str(len(contrast_names)), "All pairwise combinations")

    s("2. DATA INPUT (Peaks export)")
    kv("File format", "protein-peptides.csv", "Combined protein + peptide export")
    kv("Quantification", "Area (raw)", "Peak area under the curve (no label)")
    kv("Analysis unit", "Peptide precursor", "Sequence + Charge + Modifications = unique feature")
    kv("Score filter", f"-10lgP >= {params.get('peaks_score_thr', 20)}",
       "Peaks identification score threshold applied before analysis")
    kv("Contaminant removal", "Manual / none", "Apply upstream in Peaks if needed")

    s("3. PREPROCESSING")
    kv("Transformation", "log2(Area)", "Zeros replaced by NA")
    kv("Normalization", "Median normalization", "Per-run shift to common grand median "
       "(normalizeMedianValues equivalent); corrects global run-to-run offsets "
       "without assuming a specific distribution")
    kv("Filtering", ">= (n_rep - 1) values/condition",
       "Keep if detected in all-but-1 replicate of at least one condition")
    kv("Peptides retained", f"{n_filt_pep} / {n_raw_pep}", "After score + missing-value filter")
    kv("Imputation", "QRILC (MNAR)",
       "Quantile Regression Imputation of Left-Censored data (Lazar 2016). "
       "Suited to DDA (MNAR dominant: peptides under detection threshold are missing). "
       "Draws from a truncated normal below the estimated detection limit.")

    s("4. DIFFERENTIAL ANALYSIS")
    kv("Model", "limma eBayes",
       "Linear regression + empirical Bayes moderation (Smyth 2004). "
       "Applied directly at peptide level (no protein rollup). "
       "DEqMS not used (redundant: analysis is already at peptide level).")
    kv("Design matrix", "~0 + condition", "One-hot encoding")
    kv("Contrasts", "All pairs", "makeContrasts equivalent")
    kv("FDR correction", "Benjamini-Hochberg",
       "Global" if params.get("fdr_global") else "Per contrast")
    kv("Significance threshold", p_volc, f"= {params['volcano_p_thresh']}")
    kv("FC threshold", f"ratio >= {params['volcano_ratio_min']}",
       f"|log2FC| >= {params['volcano_lfc_min']:.3f}")
    kv("Pi-score", "|log2FC| x -log10(p)", "Combined magnitude + significance")

    s("5. ROBUSTNESS SCORE")
    kv("Iterations", str(n_iter), "MinProb re-imputations")
    kv("Criterion", f"{p_volc} < {params['volcano_p_thresh']} & ratio >= {params['volcano_ratio_min']}",
       "Identical to volcano thresholds — stability of significance under imputation uncertainty")

    s("6. ANOVA & HEATMAPS")
    kv("Test", "One-way ANOVA (F-test)", "Per peptide, on the condition factor")
    kv("FDR", "Benjamini-Hochberg", "Per-contrast" if not params.get("fdr_global") else "Global")
    kv("Threshold", p_anova, f"= {params['anova_p_thresh']}")
    kv("Significant peptides", str(n_sig_anova), "Retained for heatmaps")
    kv("Standardization", "Row Z-score", "Per-peptide centering-scaling")
    kv("Clustering", "K-means", f"n_clusters = {params.get('n_heatmap_clusters', 3)}")

    s("7. EXPLORATORY ANALYSES")
    kv("PCA", "Peptide-level PCA", "Standardized, 95% confidence ellipses")
    kv("UMAP", "n_neighbors=5, min_dist=0.1", "Non-linear, seed=42")
    kv("Correlation", "Pearson between samples", "QC replicate consistency")
    kv("UpSet", "DEP peptide intersections", "Shared across contrasts")

    s("8. PROTEIN AGGREGATION (post-hoc)")
    kv("Method", "Median per protein × sample", "Summarizes peptide-level log2 intensities; "
       "purely for overview visualisation — not used for statistical testing")

    _ptm = params.get("ptm_keys")
    if _ptm:
        s("9. PTM-TARGETED SUB-ANALYSES")
        kv("Principle", "Per-PTM differential analysis",
           "The full pipeline (volcano/ANOVA/heatmap/UpSet) is re-run on the "
           "subset of peptides carrying each selected modification")
        kv("Detection", "Text OR inline mass (combined)",
           "A peptide carries a PTM if matched by the PTM column annotation OR "
           "by an inline delta mass (±0.008-0.01 Da)")
        kv("Disambiguation", "Acetyl (+42.011) vs Trimethyl (+42.047)",
           "Tight tolerance + text annotation precedence (0.036 Da apart)")
        kv("Selected PTMs", ", ".join(str(p) for p in _ptm),
           "Independent subset analysis per modification")


def export_excel(df_raw: pd.DataFrame, df_results: pd.DataFrame,
                 df_anova: pd.DataFrame, df_intersections: pd.DataFrame,
                 mat_zscore, sig_ids, meta: pd.DataFrame,
                 contrast_names: list, design: pd.DataFrame, params: dict,
                 qc_files: list, pca_files: list, scatter_files: list,
                 volc_files: list, facet_volc: str, umap_file: str,
                 hm_corr: str, hm_classic: str, hm_clusters: str,
                 hm_violin: str, upset_file: str,
                 mat_imp: pd.DataFrame, umap_coords: pd.DataFrame,
                 n_iter: int, out_path: str,
                 prot_mat: pd.DataFrame = None,
                 ptm_results: dict = None):
    """Classeur Excel multi-onglets, structure parallèle à deimos.py."""
    print("\n[EXCEL] Building Excel report...")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    hdr_fill = PatternFill("solid", fgColor="D9D9D9")
    hdr_font = Font(bold=True)

    def write_df(ws, df, fill=hdr_fill):
        ws.append(list(df.columns))
        for cell in ws[1]:
            cell.fill = fill; cell.font = hdr_font
        for row in df.itertuples(index=False):
            ws.append(list(row))

    from openpyxl.utils import get_column_letter
    _resized_cache = {}

    def _resize(path, max_w, max_h):
        try:
            from PIL import Image as _PIL
            key = (str(path), max_w, max_h)
            if key in _resized_cache:
                return _resized_cache[key]
            im = _PIL.open(path); w, h = im.size
            scale = min(max_w/w, max_h/h, 1.0)
            if scale < 1.0:
                im = im.resize((int(w*scale), int(h*scale)), _PIL.LANCZOS)
                base, ext = os.path.splitext(str(path))
                out = f"{base}_xl{max_w}x{max_h}{ext}"
                im.save(out); _resized_cache[key] = out; return out
        except Exception:
            pass
        _resized_cache[(str(path), max_w, max_h)] = path
        return path

    def insert_img(ws, path, anchor=None, *, row=None, col=None,
                   max_w=560, max_h=440):
        if not path or not os.path.exists(str(path)):
            return
        rpath = _resize(path, max_w, max_h)
        img = XLImage(str(rpath))
        if anchor is None:
            anchor = f"{get_column_letter(col)}{row}"
        ws.add_image(img, anchor)

    IMG_COL = 10; IMG_ROW = 24

    # Methods
    ws = wb.create_sheet("Methods")
    build_methods_sheet(ws, params, n_raw_pep=len(df_raw),
                        n_filt_pep=len(mat_imp),
                        n_sig_anova=len(sig_ids) if sig_ids is not None else 0,
                        contrast_names=contrast_names,
                        design=design, n_iter=n_iter)

    # Raw data
    ws = wb.create_sheet("raw_peptides")
    write_df(ws, df_raw)

    # Log2 imputed matrix
    ws = wb.create_sheet("Log2_Impute")
    df_imp = mat_imp.copy()
    df_imp.columns = [str(c) for c in df_imp.columns]
    df_imp.insert(0, "peptide_id", meta["peptide_id"].values[:len(df_imp)])
    if "Accession" in meta.columns:
        df_imp.insert(1, "Accession", meta["Accession"].values[:len(df_imp)])
    write_df(ws, df_imp.round(4))

    # QC
    ws = wb.create_sheet("QC")
    for i, f in enumerate(qc_files):
        insert_img(ws, f, row=(i//3)*IMG_ROW+1, col=(i%3)*IMG_COL+1)

    # PCA + UMAP + corrélation
    ws = wb.create_sheet("PCA_UMAP")
    for i, f in enumerate(list(pca_files) + [umap_file, hm_corr]):
        insert_img(ws, f, row=(i//2)*IMG_ROW+1, col=(i%2)*IMG_COL+1)

    # Coordonnées UMAP
    if umap_coords is not None:
        ws = wb.create_sheet("UMAP")
        write_df(ws, umap_coords)

    # Scatter
    ws = wb.create_sheet("Scatter_Plots")
    for i, f in enumerate(scatter_files):
        insert_img(ws, f, row=(i//2)*IMG_ROW+1, col=(i%2)*IMG_COL+1)

    # Analyse différentielle
    ws = wb.create_sheet("Differential_Expression")
    cols_keep = ["peptide_id", "Sequence", "Modifications", "Charge",
                 "Accession", "Gene", "Description", "Score_10lgP"]
    for cname in contrast_names:
        for suffix in ("_diff", "_p.val", "_p.adj",
                       f"_p.val", f"_p.adj"):
            col = f"{cname}{suffix}"
            if col in df_results.columns and col not in cols_keep:
                cols_keep.append(col)
        for col in (f"Pi_Score_{cname}", f"Robustness_Score_{cname}"):
            if col in df_results.columns:
                cols_keep.append(col)
    cols_keep = [c for c in cols_keep if c in df_results.columns]
    write_df(ws, df_results[cols_keep])

    # Volcanos
    ws = wb.create_sheet("Volcano_Plots")
    ws.append([f"Facet Volcano — {len(contrast_names)} comparisons"])
    insert_img(ws, facet_volc, anchor="A3", max_w=1100, max_h=1400)
    base_row = max(60, ceil(len(contrast_names)/4)*22+8)
    for i, f in enumerate(volc_files):
        insert_img(ws, f, row=base_row+(i//3)*IMG_ROW, col=(i%3)*IMG_COL+1)

    # UpSet
    ws = wb.create_sheet("UpSet_Intersections")
    insert_img(ws, upset_file, row=1, col=1, max_w=1100, max_h=850)
    if df_intersections is not None and len(df_intersections) > 0:
        start_col = 19
        for j, colname in enumerate(df_intersections.columns):
            cell = ws.cell(row=1, column=start_col+j, value=colname)
            cell.fill = hdr_fill; cell.font = hdr_font
        for i, row in enumerate(df_intersections.itertuples(index=False), start=2):
            for j, val in enumerate(row):
                ws.cell(row=i, column=start_col+j, value=val)

    # Z-score (valeurs uniquement — la heatmap est sur l'onglet ANOVA)
    ws = wb.create_sheet("Z-score")
    if mat_zscore is not None and len(sig_ids) > 0:
        df_z = pd.DataFrame(mat_zscore, columns=design["label"].values)
        df_z.insert(0, "peptide_id", sig_ids[:len(df_z)])
        write_df(ws, df_z, fill=PatternFill("solid", fgColor="FFCC00"))

    # ANOVA
    ws = wb.create_sheet("ANOVA_Results")
    write_df(ws, df_anova)
    insert_img(ws, hm_classic, row=2, col=df_anova.shape[1]+2, max_w=900, max_h=1100)

    # ANOVA clusters
    ws = wb.create_sheet("ANOVA_Clusters")
    write_df(ws, df_anova)
    base_c = df_anova.shape[1] + 2
    insert_img(ws, hm_clusters, row=2, col=base_c, max_w=900, max_h=1100)
    insert_img(ws, hm_violin, row=2, col=base_c+IMG_COL+5, max_w=900, max_h=600)

    # Agrégation protéine (si disponible)
    if prot_mat is not None:
        ws = wb.create_sheet("Protein_Aggregation")
        write_df(ws, prot_mat)
        f_agg = os.path.join(os.path.dirname(out_path), "protein_aggregation_summary.png")
        insert_img(ws, f_agg, row=2, col=prot_mat.shape[1]+2)

    # ─── Onglets PTM-ciblés (un groupe par PTM analysé) ───
    if ptm_results:
        for key, b in ptm_results.items():
            label = b.get("label", key)
            cons  = b.get("contrasts", [])
            # 1) Onglet Differential Expression du PTM
            ws = wb.create_sheet(f"PTM_{key}_DE"[:31])
            ws.append([f"{label} — {b.get('n_peptides', 0)} peptides "
                       f"| {len(cons)} contrasts"])
            for c in ws[1]:
                c.font = Font(bold=True, size=11, color="7B2D8E")
            dfr = b.get("df_results")
            if dfr is not None:
                cols = ["peptide_id", "Sequence", "Modifications", "Charge",
                        "Accession", "Gene", "Description"]
                for c in cons:
                    cols += [f"{c}_diff", f"{c}_p.val", f"{c}_p.adj",
                             f"Pi_Score_{c}", f"Robustness_Score_{c}"]
                cols = [c for c in cols if c in dfr.columns]
                ws.append(cols)
                for cell in ws[2]:
                    cell.fill = hdr_fill; cell.font = hdr_font
                for row in dfr[cols].itertuples(index=False):
                    ws.append(list(row))
                # Facet volcano du PTM à droite du tableau
                insert_img(ws, b.get("facet_volc"),
                           row=2, col=len(cols)+2, max_w=900, max_h=1100)

            # 2) Onglet ANOVA + heatmap du PTM
            ws2 = wb.create_sheet(f"PTM_{key}_ANOVA"[:31])
            dfa = b.get("df_anova")
            if dfa is not None and len(dfa) > 0:
                write_df(ws2, dfa)
                insert_img(ws2, b.get("hm_clusters"),
                           row=2, col=dfa.shape[1]+2, max_w=850, max_h=1100)
                insert_img(ws2, b.get("hm_violin"),
                           row=2, col=dfa.shape[1]+IMG_COL+4, max_w=850, max_h=600)
            else:
                ws2.append([f"{label}: no ANOVA-significant peptides."])
        print(f"  [OK] {len(ptm_results)} PTM sub-analysis tab group(s) added.")

    wb.save(out_path)
    print(f"  [OK] Excel saved: {out_path}")


# ==============================================================================
# HELPERS
# ==============================================================================

def _condition_colors(conditions) -> dict:
    unique = list(dict.fromkeys(conditions))
    n = max(len(unique), 1)
    try:
        cmap = matplotlib.colormaps["tab20"].resampled(n)
    except AttributeError:
        cmap = plt.cm.get_cmap("tab20", n)
    return {c: matplotlib.colors.to_hex(cmap(i)) for i, c in enumerate(unique)}


def _makedirs(p):
    os.makedirs(p, exist_ok=True)
    return p


def _col_letter(n: int) -> str:
    result = ""
    while n > 0:
        n, rem = divmod(n-1, 26)
        result = chr(65+rem) + result
    return result


# ==============================================================================
# CONFIG — Adaptation de resolve_config pour Phobos
# ==============================================================================

def _resolve_phobos_config(ask_params_fn) -> dict:
    """
    Wrapper minimal autour de resolve_config de deimos/config.py.
    Phobos n'a pas de GO ni de dashboard pour l'instant ; on complète
    les clés manquantes après l'appel à resolve_config.
    """
    params = resolve_config(
        ask_params_fn=ask_params_fn,
        ask_go_params_fn=lambda: None,
        go_available=False,
        dash_available=False,
        # config.py (Deimos) fait os.path.exists(pr_path) sans garde None :
        # on passe une chaîne vide (Phobos n'utilise pas de pr_matrix/DEqMS).
        pr_path="",
    )
    # Valeurs par défaut Phobos (si absentes du YAML)
    params.setdefault("peaks_csv",     "protein-peptides.csv")
    params.setdefault("peaks_score_thr", 20.0)
    params.setdefault("impute_method", "qrilc")
    params.setdefault("make_dashboard", True)
    params.setdefault("use_deqms",     False)
    params.setdefault("go_organism",   None)
    # fasta_path : None = auto-détection d'un .fasta dans le dossier d'entrée
    params.setdefault("fasta_path",    None)
    # ptm_keys : None = sélection interactive ; [] = aucune ; liste = clés PTM
    # (Phospho, Oxidation, Methyl_all, Acetyl, Dimethyl, Trimethyl, Methyl)
    if "ptm_keys" not in params:
        params["ptm_keys"] = None
    return params


# ==============================================================================
# MAIN
# ==============================================================================

def main():
    params = _resolve_phobos_config(ask_params_fn=ask_params)

    csv_path    = params.get("peaks_csv",   params.get("tsv_path", "protein-peptides.csv"))
    design_path = params["design_path"]
    out_dir     = params["out_dir"]
    score_thr   = float(params.get("peaks_score_thr", 20.0))
    N_ITER      = params.get("n_iter_robustness", 100)

    _makedirs(out_dir)

    # 0. Diagnostic
    if not diagnose_labels_peaks(csv_path, design_path):
        sys.exit(1)

    # 1. Chargement + parsing Peaks
    print("\n[STEP] Loading Peaks protein-peptides.csv...")
    df_raw, meta, mat_area, design, area_cols = load_peaks_data(
        csv_path, design_path, score_thr=score_thr)

    # 1bis. Récupération des descriptions protéiques depuis un FASTA (auto)
    # Si un .fasta est présent dans le dossier du fichier d'entrée, on complète
    # la colonne Description (jointure sur l'Accession, 1er ID avant '|').
    desc_missing = ("Description" not in meta.columns or
                    meta.get("Description", pd.Series(dtype=str))
                        .fillna("").astype(str).str.strip().eq("").all())
    if desc_missing:
        try:
            from fasta_descriptions import add_descriptions
            search_dir = os.path.dirname(os.path.abspath(csv_path)) or "."
            meta, fasta_info = add_descriptions(
                meta, fasta_path=params.get("fasta_path"),
                search_dir=search_dir, accession_col="Accession")
            if not fasta_info["used"]:
                print("  [INFO] No FASTA found — descriptions left empty "
                      "(drop a .fasta in the folder to enable recovery).")
        except ImportError:
            print("  [INFO] fasta_descriptions module not found — skipped.")
    
    # 2. Log2 + normalisation médiane
    print("\n[STEP] log2 + median normalization...")
    mat_log2 = log2_normalize(mat_area)
    mat_log2.columns = design["label"].values
    mat_area.columns = design["label"].values

    # 3. QC pré-filtrage
    print("\n[QC] Quality control (pre-filter)...")
    qc_files = plot_qc(mat_log2, design, out_dir)

    # 4. Filtrage
    print("\n[STEP] Filtering...")
    mat_filt, keep_mask = filter_missval(mat_log2, design, thr=1)
    meta_filt = meta[keep_mask].reset_index(drop=True)

    # Diagnostic missingness
    miss_diag = diagnose_missingness(mat_filt, design, out_dir)
    if miss_diag.get("plot"):
        qc_files.append(miss_diag["plot"])

    # 5. Imputation QRILC
    print("\n[STEP] QRILC imputation...")
    mat_imp = impute_qrilc(mat_filt)
    mat_imp.columns = design["label"].values
    mat_filt.columns = design["label"].values
    qc_imp = plot_imputation(mat_filt, mat_imp, out_dir)
    qc_files.append(qc_imp)

    # 6. Analyse différentielle + robustness
    print("\n[STEP] Differential analysis...")
    df_results, contrast_names, scatter_files = run_differential_analysis(
        mat_imp, mat_filt, meta_filt, design, params, out_dir, n_iter=N_ITER)

    # 7. Volcanos
    print("\n[STEP] Volcano plots...")
    volc_files, facet_volc = plot_volcanoes(df_results, contrast_names, params, out_dir)

    # 8. PCA
    print("\n[STEP] PCA...")
    pca_files = plot_pca(mat_imp, design, out_dir)

    # 9. Corrélation
    print("\n[STEP] Correlation heatmap...")
    hm_corr = plot_correlation_heatmap(mat_imp, design, out_dir)

    # 10. UMAP
    print("\n[STEP] UMAP...")
    umap_file, umap_coords = plot_umap(mat_imp, design, out_dir)

    # 11. ANOVA + heatmaps
    print("\n[STEP] ANOVA & heatmaps...")
    (df_anova, hm_classic, hm_clusters, hm_violin,
     mat_zscore, cluster_mapping, sig_ids) = run_anova_heatmaps(
        mat_imp, meta_filt, design, params, out_dir)

    # 12. UpSet
    print("\n[STEP] UpSet plot...")
    upset_file, df_intersections = plot_upset(df_results, contrast_names, params, out_dir)

    # 13. Agrégation protéine (post-hoc, optionnelle)
    prot_mat = aggregate_to_protein(df_results, mat_imp, design,
                                    contrast_names, params, out_dir)

    # 13bis. Sous-analyses PTM-ciblées (optionnel, non bloquant)
    ptm_results = {}
    ptm_keys = params.get("ptm_keys", None)
    if ptm_keys is None:
        # Sélection interactive si non fournie par la config YAML
        try:
            from ptm_analysis import ask_ptm_selection
            ptm_keys = ask_ptm_selection(meta_filt)
        except Exception as e:
            print(f"  [WARN] PTM selection unavailable: {e}")
            ptm_keys = []
    if ptm_keys:
        try:
            from ptm_analysis import run_ptm_subanalyses
            callbacks = {
                "differential": run_differential_analysis,
                "volcanoes":    plot_volcanoes,
                "anova":        run_anova_heatmaps,
                "upset":        plot_upset,
            }
            ptm_results = run_ptm_subanalyses(
                ptm_keys, mat_filt, mat_imp, meta_filt, design, params,
                out_dir, callbacks, n_iter=N_ITER)
        except Exception as e:
            print(f"  [WARN] PTM sub-analyses failed: {type(e).__name__}: {e}")
            ptm_results = {}

    # 14. Export Excel
    out_name = os.path.join(out_dir, "PeptideAnalysis_Results.xlsx")
    export_excel(
        df_raw=df_raw, df_results=df_results, df_anova=df_anova,
        df_intersections=df_intersections, mat_zscore=mat_zscore,
        sig_ids=sig_ids, meta=meta_filt, contrast_names=contrast_names,
        design=design, params=params,
        qc_files=qc_files, pca_files=pca_files, scatter_files=scatter_files,
        volc_files=volc_files, facet_volc=facet_volc, umap_file=umap_file,
        hm_corr=hm_corr, hm_classic=hm_classic, hm_clusters=hm_clusters,
        hm_violin=hm_violin, upset_file=upset_file,
        mat_imp=mat_imp, umap_coords=umap_coords,
        n_iter=N_ITER, out_path=out_name, prot_mat=prot_mat,
        ptm_results=ptm_results
    )

    # 15. Dashboard HTML interactif (optionnel, non bloquant)
    if params.get("make_dashboard", False):
        print("\n[STEP] Building interactive HTML dashboard...")
        try:
            from build_dashboard_phobos import build_dashboard
            dash_out = os.path.join(out_dir, "phobos_dashboard.html")
            result = build_dashboard(out_name, design, dash_out, params=params)
            if result:
                print(f"  [OK] Dashboard: {result}")
            else:
                print("  [WARN] Dashboard not generated (see messages above).")
        except ImportError:
            print("  [WARN] build_dashboard_phobos.py not found — dashboard skipped.")
            print("         Place it alongside phobos.py to enable the dashboard.")
    else:
        print("\n[SKIP] Dashboard not requested (make_dashboard=false).")

    # 16. Nettoyage figures temporaires (volcanos conservés)
    print("\n[CLEAN] Removing temporary figures...")
    removed = 0
    def _clean_dir(d):
        nonlocal removed
        for fn in os.listdir(d):
            full = os.path.join(d, fn)
            if os.path.isdir(full) and fn.startswith("ptm_"):
                _clean_dir(full)          # nettoyer aussi les sous-dossiers PTM
                continue
            if not fn.lower().endswith((".png", ".tiff", ".tif")):
                continue
            is_volc = fn.startswith("volc_")
            is_resized = "_xl" in fn and "x" in fn.split("_xl")[-1]
            if is_volc and not is_resized:
                continue
            try:
                os.remove(full); removed += 1
            except OSError:
                pass
    _clean_dir(out_dir)
    print(f"  [OK] {removed} temp figures removed (volcanos kept)")

    print(f"\n[DONE] Phobos pipeline complete.")
    print(f"   -> {out_name}")


if __name__ == "__main__":
    main()
